from __future__ import annotations

import csv
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable
import re
import io

from django.utils import timezone

from prices import models


@dataclass
class ParsedRow:
    sku: str
    name: str
    price: Decimal
    currency: str


def _parse_decimal(value) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        # Keep true numeric spreadsheet values (e.g. 13.1999) as-is.
        # Heuristic string parsing below is for localized text input.
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    # Keep only number-ish chunk, including common thousands separators.
    match = re.search(r"[-+]?\d[\d\s\u00A0\u202F.,']*", text)
    if not match:
        return None
    text = match.group(0)
    # Normalize all kinds of spaces and apostrophe thousands separators.
    text = re.sub(r"[\s\u00A0\u202F']", "", text)
    if not text:
        return None

    # Resolve decimal separator heuristically.
    has_comma = "," in text
    has_dot = "." in text
    if has_comma and has_dot:
        # Last separator is decimal, other one is thousands.
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "")
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
    elif has_comma:
        parts = text.split(",")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            text = parts[0] + "." + parts[1]
        elif len(parts) >= 2 and len(parts[-1]) in (1, 2):
            text = "".join(parts[:-1]) + "." + parts[-1]
        else:
            text = "".join(parts)
    elif has_dot:
        parts = text.split(".")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            text = parts[0] + "." + parts[1]
        elif len(parts) >= 2 and len(parts[-1]) in (1, 2):
            text = "".join(parts[:-1]) + "." + parts[-1]
        else:
            text = "".join(parts)
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _detect_currency(cell_value) -> str | None:
    if cell_value is None:
        return None
    text = _fix_mojibake(str(cell_value).strip())
    if not text:
        return None
    upper = text.upper()
    lower = text.lower()
    rub_upper = "\u0420\u0423\u0411"
    rub_lower = "\u0440\u0443\u0431"
    dol_lower = "\u0434\u043e\u043b"
    if "USD" in upper or "$" in upper or dol_lower in lower:
        return models.Currency.USD
    if (
        "RUB" in upper
        or "RUR" in upper
        or rub_upper in upper
        or rub_lower in lower
        or "Ð Ð£Ð‘" in upper
        or "â‚½" in upper
        or "\u20BD" in text
    ):
        return models.Currency.RUB
    return None


def _count_cyrillic(text: str) -> int:
    return len(re.findall(r"[\u0400-\u04FF]", text))


def _fix_mojibake(text: str) -> str:
    if not text:
        return text
    if _count_cyrillic(text):
        return text
    if not re.search(r"[\u00C0-\u00FF]", text):
        return text
    best = text
    best_score = _count_cyrillic(text)
    for encoding in ("cp1251", "utf-8"):
        try:
            candidate = text.encode("latin-1").decode(encoding)
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
        score = _count_cyrillic(candidate)
        if score > best_score:
            best = candidate
            best_score = score
    return best


def _read_csv_text_from_bytes(raw: bytes) -> str:
    if not raw:
        return ""
    best_text = ""
    best_score = -1
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            decoded = raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        score = _count_cyrillic(decoded)
        if score > best_score:
            best_score = score
            best_text = decoded
    if best_text:
        return best_text
    return raw.decode("utf-8-sig", errors="ignore")


def _normalize_sku(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, Decimal):
        if value == value.to_integral():
            return str(int(value))
        return str(value.normalize())
    text = str(value).strip()
    if not text:
        return ""
    if re.match(r"^-?\d+\.0+$", text):
        return text.split(".")[0]
    return text


def _legacy_sku_variants(sku: str) -> list[str]:
    if not sku:
        return []
    variants = []
    if re.match(r"^-?\d+$", sku):
        variants.append(f"{sku}.0")
    if re.match(r"^-?\d+\.0+$", sku):
        variants.append(sku.split(".")[0])
    return variants


def _iter_rows_csv(path: Path, start_row: int) -> Iterable[list[str]]:
    raw = path.read_bytes()
    text = _read_csv_text_from_bytes(raw)
    reader = csv.reader(io.StringIO(text))
    for index, row in enumerate(reader, start=1):
        if index < start_row:
            continue
        yield row


def _iter_rows_csv_file(file_obj, start_row: int) -> Iterable[list[str]]:
    file_obj.seek(0)
    raw = file_obj.read()
    text = _read_csv_text_from_bytes(raw)
    reader = csv.reader(io.StringIO(text))
    for index, row in enumerate(reader, start=1):
        if index < start_row:
            continue
        yield row


def _iter_rows_xlsx(
    path: Path, sheet_name: str, sheet_index: int | None, start_row: int
):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx files") from exc

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name:
            sheet = workbook[sheet_name]
        elif sheet_index is not None:
            sheet = workbook.worksheets[sheet_index]
        else:
            sheet = workbook.active

        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            yield list(row)
    finally:
        workbook.close()


def _iter_rows_xlsx_sheet(
    workbook, sheet_name: str, sheet_index: int | None, start_row: int
):
    if sheet_name:
        sheet = workbook[sheet_name]
    elif sheet_index is not None:
        sheet = workbook.worksheets[sheet_index]
    else:
        sheet = workbook.active
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        yield list(row)


def _iter_rows_xlsx_file(file_obj, sheet_name: str, sheet_index: int | None, start_row: int):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx files") from exc

    file_obj.seek(0)
    workbook = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    try:
        if sheet_name:
            sheet = workbook[sheet_name]
        elif sheet_index is not None:
            sheet = workbook.worksheets[sheet_index]
        else:
            sheet = workbook.active

        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            yield list(row)
    finally:
        workbook.close()


def _iter_rows_xls_path(path: Path, sheet_name: str, sheet_index: int | None, start_row: int):
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("xlrd is required to read .xls files") from exc

    workbook = xlrd.open_workbook(path)
    if sheet_name:
        sheet = workbook.sheet_by_name(sheet_name)
    elif sheet_index is not None:
        sheet = workbook.sheet_by_index(sheet_index)
    else:
        sheet = workbook.sheet_by_index(0)
    for row_idx in range(start_row - 1, sheet.nrows):
        yield sheet.row_values(row_idx)


def _iter_rows_xls_sheet(sheet, start_row: int):
    for row_idx in range(start_row - 1, sheet.nrows):
        yield sheet.row_values(row_idx)


def _iter_rows_xls_file(file_obj, sheet_name: str, sheet_index: int | None, start_row: int):
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("xlrd is required to read .xls files") from exc

    file_obj.seek(0)
    workbook = xlrd.open_workbook(file_contents=file_obj.read())
    if sheet_name:
        sheet = workbook.sheet_by_name(sheet_name)
    elif sheet_index is not None:
        sheet = workbook.sheet_by_index(sheet_index)
    else:
        sheet = workbook.sheet_by_index(0)
    for row_idx in range(start_row - 1, sheet.nrows):
        yield sheet.row_values(row_idx)


def _parse_rows(
    rows: Iterable[list],
    sku_col: int,
    name_cols: list[int],
    price_col: int,
    currency_col: int,
    default_currency: str,
    skip_until_valid: bool = True,
    parse_stats: dict | None = None,
) -> Iterable[ParsedRow]:
    sku_idx = sku_col - 1 if sku_col else None
    name_indexes = [value - 1 for value in name_cols if value]
    price_idx = price_col - 1
    currency_idx = currency_col - 1 if currency_col else None
    found_data = not skip_until_valid
    skip_terms = ("итого", "итог", "доставка")
    for row in rows:
        max_index = max([price_idx, *name_indexes, sku_idx if sku_idx is not None else 0])
        if len(row) <= max_index:
            if parse_stats is not None:
                parse_stats["row_too_short"] = parse_stats.get("row_too_short", 0) + 1
            continue
        sku = ""
        if sku_idx is not None and sku_idx < len(row):
            sku = _normalize_sku(row[sku_idx])
        name_parts = []
        for index in name_indexes:
            value = row[index] if index < len(row) else None
            text = str(value).strip() if value is not None else ""
            text = _fix_mojibake(text).strip()
            if text and not re.match(r"^-?\d+(?:[.,]\d+)?$", text):
                name_parts.append(text)
        name = " ".join(name_parts).strip()
        if name:
            lowered = name.lower()
            if any(term in lowered for term in skip_terms):
                if parse_stats is not None:
                    parse_stats["skip_term"] = parse_stats.get("skip_term", 0) + 1
                if not found_data:
                    continue
                continue
            if _is_invalid_short_name(name):
                if parse_stats is not None:
                    parse_stats["invalid_short_name"] = parse_stats.get("invalid_short_name", 0) + 1
                if not found_data:
                    continue
                continue
        price = _parse_decimal(row[price_idx])
        if not name or price is None or price == 0:
            if parse_stats is not None:
                if not name:
                    parse_stats["empty_name"] = parse_stats.get("empty_name", 0) + 1
                if price is None:
                    parse_stats["invalid_price"] = parse_stats.get("invalid_price", 0) + 1
                elif price == 0:
                    parse_stats["zero_price"] = parse_stats.get("zero_price", 0) + 1
            if not found_data:
                continue
            continue
        detected_currency = None
        if currency_idx is not None and currency_idx < len(row):
            detected_currency = _detect_currency(row[currency_idx])
        if not detected_currency:
            detected_currency = _detect_currency(row[price_idx])
        detected_currency = detected_currency or default_currency
        found_data = True
        if parse_stats is not None:
            parse_stats["parsed"] = parse_stats.get("parsed", 0) + 1
        yield ParsedRow(sku=sku, name=name, price=price, currency=detected_currency)


def _identity_key(sku: str, name: str) -> str:
    if sku:
        return _normalize_sku(sku)
    return re.sub(r"\s+", " ", name.strip()).lower()


def _is_invalid_short_name(name: str) -> bool:
    normalized = re.sub(r"\s+", " ", (name or "").strip())
    if len(normalized) < 3:
        return True
    words = [word for word in normalized.split(" ") if word]
    if len(words) == 1 and len(words[0]) < 10:
        return True
    return False


def _get_historical_usd_rub_rate(rate_date):
    rate = (
        models.ExchangeRate.objects.filter(
            from_currency=models.Currency.USD,
            to_currency=models.Currency.RUB,
            rate_date__lte=rate_date,
        )
        .order_by("-rate_date", "-id")
        .first()
    )
    if rate:
        return rate.rate
    return None


def _get_rates_for_date(rate_date):
    rates = {}
    for rate in models.ExchangeRate.objects.filter(rate_date__lte=rate_date).order_by(
        "-rate_date", "-id"
    ):
        key = (rate.from_currency, rate.to_currency)
        if key not in rates:
            rates[key] = rate.rate
    return rates


def _convert_price_with_rates(
    amount: Decimal,
    from_currency: str,
    to_currency: str,
    rates: dict[tuple[str, str], Decimal],
) -> Decimal | None:
    if amount is None or not from_currency or not to_currency:
        return None
    if from_currency == to_currency:
        return amount
    direct = rates.get((from_currency, to_currency))
    if direct and direct != 0:
        return (amount * direct).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    inverse = rates.get((to_currency, from_currency))
    if inverse and inverse != 0:
        return (amount / inverse).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return None


def _compute_snapshot_prices(price: Decimal, currency: str, usd_rub_rate):
    amount = Decimal(price)
    if currency == models.Currency.USD:
        usd_value = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        rub_value = None
        if usd_rub_rate:
            rub_value = (amount * usd_rub_rate).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
        return rub_value, usd_value
    if currency == models.Currency.RUB:
        rub_value = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        usd_value = None
        if usd_rub_rate and usd_rub_rate != 0:
            usd_value = (amount / usd_rub_rate).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
        return rub_value, usd_value
    return None, None


def process_import_file(import_file: models.ImportFile) -> None:
    if not import_file.file:
        raise RuntimeError("Import file is missing.")

    mapping = import_file.mapping
    if not mapping:
        raise RuntimeError("Mapping is missing.")

    if import_file.file_kind != models.FileKind.PRICE:
        import_file.status = models.ImportStatus.PROCESSED
        import_file.processed_at = timezone.now()
        import_file.save(update_fields=["status", "processed_at"])
        return

    path = Path(import_file.file.path)
    extension = path.suffix.lower()
    start_row = mapping.header_row or 1
    column_map = mapping.column_map or {}
    sku_col = int(column_map.get("sku", 0))
    name_value = column_map.get("name", 0)
    if isinstance(name_value, list):
        name_cols = [int(value) for value in name_value if value]
    else:
        name_cols = [int(name_value)] if name_value else []
    price_col = int(column_map.get("price", 0))
    currency_col = int(column_map.get("currency", 0) or 0)
    if not name_cols or not price_col:
        raise RuntimeError("Mapping must include name and price columns.")
    parse_stats: dict[str, int] = {}

    def _collect_parsed(rows_iter):
        return list(
            _parse_rows(
                rows_iter,
                sku_col,
                name_cols,
                price_col,
                currency_col,
                import_file.import_batch.supplier.default_currency,
                skip_until_valid=True,
                parse_stats=parse_stats,
            )
        )

    if extension in {".csv"}:
        rows = _iter_rows_csv(path, start_row)
        parsed_rows = _collect_parsed(rows)
    elif extension in {".xlsx"}:
        parsed_rows = []
        sheet_names = [
            name.strip()
            for name in (mapping.sheet_names or "").split(",")
            if name.strip()
        ]
        sheet_indexes = [
            int(index.strip())
            for index in (mapping.sheet_indexes or "").split(",")
            if index.strip().isdigit()
        ]
        if sheet_names or sheet_indexes:
            for name in sheet_names:
                rows = _iter_rows_xlsx(path, name, None, start_row)
                parsed_rows.extend(_collect_parsed(rows))
            for index in sheet_indexes:
                rows = _iter_rows_xlsx(path, "", index, start_row)
                parsed_rows.extend(_collect_parsed(rows))
        else:
            try:
                import openpyxl
            except ImportError as exc:
                raise RuntimeError("openpyxl is required to read .xlsx files") from exc

            workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
            try:
                rows = _iter_rows_xlsx_sheet(workbook, mapping.sheet_name, mapping.sheet_index, start_row)
                parsed_rows = _collect_parsed(rows)
                # Some supplier files keep the price table on a non-active sheet.
                if not parsed_rows and not mapping.sheet_name and mapping.sheet_index is None:
                    for sheet in workbook.worksheets:
                        rows = _iter_rows_xlsx_sheet(workbook, sheet.title, None, start_row)
                        candidate_rows = _collect_parsed(rows)
                        if candidate_rows:
                            parsed_rows = candidate_rows
                            break
            finally:
                workbook.close()
    elif extension in {".xls"}:
        parsed_rows = []
        sheet_names = [
            name.strip()
            for name in (mapping.sheet_names or "").split(",")
            if name.strip()
        ]
        sheet_indexes = [
            int(index.strip())
            for index in (mapping.sheet_indexes or "").split(",")
            if index.strip().isdigit()
        ]
        if sheet_names or sheet_indexes:
            for name in sheet_names:
                rows = _iter_rows_xls_path(path, name, None, start_row)
                parsed_rows.extend(_collect_parsed(rows))
            for index in sheet_indexes:
                rows = _iter_rows_xls_path(path, "", index, start_row)
                parsed_rows.extend(_collect_parsed(rows))
        else:
            try:
                import xlrd
            except ImportError as exc:
                raise RuntimeError("xlrd is required to read .xls files") from exc

            workbook = xlrd.open_workbook(path)
            if mapping.sheet_name:
                selected_sheet = workbook.sheet_by_name(mapping.sheet_name)
            elif mapping.sheet_index is not None:
                selected_sheet = workbook.sheet_by_index(mapping.sheet_index)
            else:
                selected_sheet = workbook.sheet_by_index(0)
            rows = _iter_rows_xls_sheet(selected_sheet, start_row)
            parsed_rows = _collect_parsed(rows)
            if not parsed_rows and not mapping.sheet_name and mapping.sheet_index is None:
                for idx in range(workbook.nsheets):
                    rows = _iter_rows_xls_sheet(workbook.sheet_by_index(idx), start_row)
                    candidate_rows = _collect_parsed(rows)
                    if candidate_rows:
                        parsed_rows = candidate_rows
                        break
    else:
        raise RuntimeError(f"Unsupported file type: {extension}")

    received_at = import_file.import_batch.received_at
    if received_at and timezone.is_naive(received_at):
        received_at = timezone.make_aware(received_at)
    now = received_at or timezone.now()
    rate_lookup_date = timezone.localtime(now).date() if timezone.is_aware(now) else now.date()
    usd_rub_rate = _get_historical_usd_rub_rate(rate_lookup_date)
    rates_for_date = _get_rates_for_date(rate_lookup_date)
    supplier_currency = import_file.import_batch.supplier.default_currency

    if not parsed_rows:
        # Some supplier files shift header/data row unexpectedly.
        # If the configured start row yields nothing, retry a small range.
        fallback_rows = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
        fallback_rows = [row for row in fallback_rows if row != start_row]
        for candidate_start in fallback_rows:
            try:
                if extension in {".csv"}:
                    rows = _iter_rows_csv(path, candidate_start)
                    candidate_parsed = _collect_parsed(rows)
                elif extension in {".xlsx"}:
                    rows = _iter_rows_xlsx(path, mapping.sheet_name, mapping.sheet_index, candidate_start)
                    candidate_parsed = _collect_parsed(rows)
                elif extension in {".xls"}:
                    rows = _iter_rows_xls_path(path, mapping.sheet_name, mapping.sheet_index, candidate_start)
                    candidate_parsed = _collect_parsed(rows)
                else:
                    candidate_parsed = []
            except Exception:
                candidate_parsed = []
            if candidate_parsed:
                parsed_rows = candidate_parsed
                break

    if not parsed_rows:
        required_cols = []
        if sku_col:
            required_cols.append(("sku", sku_col))
        for idx, col in enumerate(name_cols, start=1):
            required_cols.append((f"name{idx}", col))
        required_cols.append(("price", price_col))
        if currency_col:
            required_cols.append(("currency", currency_col))
        required_max_col = max((col for _, col in required_cols), default=0)

        max_cols_hint = None
        selected_sheet_hint = ""
        try:
            if extension == ".xlsx":
                import openpyxl

                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
                try:
                    if mapping.sheet_name:
                        sh = wb[mapping.sheet_name]
                    elif mapping.sheet_index is not None:
                        sh = wb.worksheets[mapping.sheet_index]
                    else:
                        sh = wb.active
                    max_cols_hint = int(sh.max_column or 0)
                    selected_sheet_hint = sh.title
                finally:
                    wb.close()
            elif extension == ".xls":
                import xlrd

                wb = xlrd.open_workbook(path)
                if mapping.sheet_name:
                    sh = wb.sheet_by_name(mapping.sheet_name)
                    selected_sheet_hint = mapping.sheet_name
                elif mapping.sheet_index is not None:
                    sh = wb.sheet_by_index(mapping.sheet_index)
                    selected_sheet_hint = sh.name
                else:
                    sh = wb.sheet_by_index(0)
                    selected_sheet_hint = sh.name
                max_cols_hint = int(sh.ncols or 0)
            elif extension == ".csv":
                rows = list(_iter_rows_csv(path, start_row))
                max_cols_hint = max((len(r) for r in rows), default=0)
        except Exception:
            max_cols_hint = None

        missing_cols = []
        if max_cols_hint is not None and max_cols_hint < required_max_col:
            for label, col in required_cols:
                if col > max_cols_hint:
                    missing_cols.append(f"{label}:{col}")

        name_cols_text = ",".join(str(col) for col in name_cols) or "-"
        missing_cols_text = ", ".join(missing_cols) if missing_cols else "-"
        detected_cols_text = str(max_cols_hint) if max_cols_hint is not None else "unknown"
        selected_sheet_text = selected_sheet_hint or "-"
        raise RuntimeError(
            "No data rows parsed. "
            f"Selected sheet: {selected_sheet_text}. "
            f"Header row: {start_row}. "
            f"Configured columns: sku={sku_col or '-'}, name={name_cols_text}, "
            f"price={price_col}, currency={currency_col or '-'}; "
            f"detected max columns in sheet: {detected_cols_text}. "
            f"Missing configured columns: {missing_cols_text}. "
            f"Row diagnostics: {parse_stats}."
        )
    unique_rows = {}
    for parsed in parsed_rows:
        key = _identity_key(parsed.sku, parsed.name)
        if not key:
            continue
        normalized_price = _convert_price_with_rates(
            parsed.price,
            parsed.currency,
            supplier_currency,
            rates_for_date,
        )
        if normalized_price is None:
            raise RuntimeError(
                f"Missing exchange rate for {parsed.currency}->{supplier_currency} "
                f"on or before {rate_lookup_date}."
            )
        unique_rows[key] = ParsedRow(
            sku=parsed.sku,
            name=parsed.name,
            price=normalized_price,
            currency=supplier_currency,
        )
    if len(unique_rows) < 100:
        raise RuntimeError(
            f"Too few products ({len(unique_rows)}). Expected at least 100."
        )

    supplier = import_file.import_batch.supplier
    batch_received = import_file.import_batch.received_at
    batch_created = import_file.import_batch.created_at
    batch_time = batch_received or batch_created or timezone.now()
    supplier_latest_batch = None
    for received_at, created_at in models.ImportBatch.objects.filter(
        supplier=supplier,
        importfile__status=models.ImportStatus.PROCESSED,
        importfile__file_kind=models.FileKind.PRICE,
    ).values_list("received_at", "created_at"):
        candidate = received_at or created_at
        if candidate and (supplier_latest_batch is None or candidate > supplier_latest_batch):
            supplier_latest_batch = candidate
    identity_keys = list(unique_rows.keys())
    existing_products = models.SupplierProduct.objects.filter(
        supplier=supplier, identity_key__in=identity_keys
    )
    existing_map = {}
    for product in existing_products:
        key = product.identity_key or _identity_key(product.supplier_sku, product.name)
        if key:
            existing_map[key] = product
        for variant in _legacy_sku_variants(product.supplier_sku):
            existing_map.setdefault(variant, product)

    to_create = []
    to_update = []
    is_latest_batch = not supplier_latest_batch or batch_time >= supplier_latest_batch
    for identity_key, parsed in unique_rows.items():
        existing = existing_map.get(identity_key)
        if existing is None:
            is_current = is_latest_batch
            to_create.append(
                models.SupplierProduct(
                    supplier=supplier,
                    supplier_sku=parsed.sku,
                    identity_key=identity_key,
                    name=parsed.name,
                    currency=parsed.currency,
                    current_price=parsed.price,
                    last_imported_at=now,
                    last_import_batch=import_file.import_batch,
                    created_import_batch=import_file.import_batch,
                    is_active=is_current,
                )
            )
        else:
            existing_batch_time = None
            if existing.last_import_batch_id:
                existing_batch_time = (
                    existing.last_import_batch.received_at
                    or existing.last_import_batch.created_at
                )
            if not existing_batch_time:
                existing_batch_time = existing.last_imported_at
            changed = (
                existing.name != parsed.name
                or existing.current_price != parsed.price
                or existing.currency != parsed.currency
                or existing.last_import_batch_id != import_file.import_batch_id
                or not existing.is_active
            )
            if changed or is_latest_batch:
                if not existing_batch_time or batch_time >= existing_batch_time:
                    if existing.supplier_sku != parsed.sku:
                        sku_taken = models.SupplierProduct.objects.filter(
                            supplier=supplier, supplier_sku=parsed.sku
                        ).exclude(id=existing.id).exists()
                        if parsed.sku and not sku_taken:
                            existing.supplier_sku = parsed.sku
                    if changed:
                        existing.name = parsed.name
                        existing.current_price = parsed.price
                        existing.currency = parsed.currency
                    existing.last_imported_at = now
                    existing.last_import_batch = import_file.import_batch
                    if is_latest_batch:
                        existing.is_active = True
                    existing.identity_key = identity_key
                    to_update.append(existing)

    if to_create:
        models.SupplierProduct.objects.bulk_create(to_create, batch_size=500)
    if to_update:
        models.SupplierProduct.objects.bulk_update(
            to_update,
            [
                "supplier_sku",
                "identity_key",
                "name",
                "current_price",
                "currency",
                "last_imported_at",
                "last_import_batch",
                "is_active",
            ],
            batch_size=500,
        )

    product_map = {
        product.identity_key: product
        for product in models.SupplierProduct.objects.filter(
            supplier=supplier, identity_key__in=identity_keys
        )
    }
    snapshots = []
    for identity_key, parsed in unique_rows.items():
        product = product_map.get(identity_key)
        if product:
            price_rub, price_usd = _compute_snapshot_prices(
                parsed.price, parsed.currency, usd_rub_rate
            )
            snapshots.append(
                models.PriceSnapshot(
                    supplier_product=product,
                    import_batch=import_file.import_batch,
                    price=parsed.price,
                    currency=parsed.currency,
                    price_rub=price_rub,
                    price_usd=price_usd,
                    recorded_at=now,
                )
            )
    if snapshots:
        models.PriceSnapshot.objects.bulk_create(snapshots, batch_size=500)

    if is_latest_batch:
        models.SupplierProduct.objects.filter(supplier=supplier).exclude(
            identity_key__in=identity_keys
        ).update(is_active=False)

    import_file.status = models.ImportStatus.PROCESSED
    import_file.processed_at = timezone.now()
    import_file.save(update_fields=["status", "processed_at"])


def delete_import_batch(import_batch: models.ImportBatch) -> None:
    products = models.SupplierProduct.objects.filter(
        created_import_batch=import_batch
    )
    legacy_products = models.SupplierProduct.objects.filter(
        created_import_batch__isnull=True, last_import_batch=import_batch
    )
    models.PriceSnapshot.objects.filter(import_batch=import_batch).delete()
    models.StockSnapshot.objects.filter(import_batch=import_batch).delete()

    for import_file in import_batch.importfile_set.all():
        if import_file.file:
            import_file.file.delete(save=False)
    import_batch.importfile_set.all().delete()
    products.delete()
    legacy_products.delete()
    import_batch.delete()


def preview_mapping_file(file_obj, sheet_index: int | None = None) -> dict:
    file_obj.seek(0)
    filename = getattr(file_obj, "name", "")
    extension = Path(filename).suffix.lower()
    rows = []
    max_cols = 0
    sheet_names = []
    start_row = 1

    def _clean_preview_cell(value) -> str:
        if value is None:
            return ""
        text = str(value)
        if not text:
            return ""
        return _fix_mojibake(text)

    if extension == ".csv":
        for row in _iter_rows_csv_file(file_obj, start_row):
            rows.append([_clean_preview_cell(cell) for cell in row])
            max_cols = max(max_cols, len(row))
            if len(rows) >= 50:
                break
    elif extension == ".xlsx":
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to read .xlsx files") from exc
        file_obj.seek(0)
        workbook = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
        try:
            sheet_names = workbook.sheetnames
            if sheet_index is not None and 0 <= sheet_index < len(workbook.worksheets):
                sheet = workbook.worksheets[sheet_index]
            else:
                sheet = workbook.active
            max_cols = sheet.max_column or 0
            for row in sheet.iter_rows(min_row=start_row, values_only=True):
                row_values = list(row)
                if max_cols and len(row_values) < max_cols:
                    row_values.extend([""] * (max_cols - len(row_values)))
                rows.append([_clean_preview_cell(cell) for cell in row_values])
                if len(rows) >= 50:
                    break
        finally:
            workbook.close()
    elif extension == ".xls":
        file_obj.seek(0)
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("xlrd is required to read .xls files") from exc
        workbook = xlrd.open_workbook(file_contents=file_obj.read())
        sheet_names = workbook.sheet_names()
        if sheet_index is not None and 0 <= sheet_index < len(sheet_names):
            sheet = workbook.sheet_by_index(sheet_index)
        else:
            sheet = workbook.sheet_by_index(0)
        max_cols = sheet.ncols
        for row_idx in range(start_row - 1, sheet.nrows):
            row = sheet.row_values(row_idx)
            if max_cols and len(row) < max_cols:
                row.extend([""] * (max_cols - len(row)))
            rows.append([_clean_preview_cell(cell) for cell in row])
            if len(rows) >= 50:
                break
    else:
        raise RuntimeError("Unsupported file type for preview.")

    col_offset = 0
    if rows and max_cols:
        for col_index in range(max_cols):
            if all((row[col_index] or "").strip() == "" for row in rows):
                col_offset += 1
            else:
                break

    return {
        "rows": rows,
        "max_cols": max_cols,
        "sheet_names": sheet_names,
        "col_offset": col_offset,
    }
