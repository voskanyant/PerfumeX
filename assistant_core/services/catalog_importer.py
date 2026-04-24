from __future__ import annotations

import csv
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import StringIO
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from assistant_linking.models import BrandAlias, ProductAlias
from assistant_linking.services.normalizer import normalize_text
from catalog.models import Brand, Perfume, PerfumeVariant


COLUMN_ALIASES = {
    "brand": {"brand", "brand_name", "brand name", "manufacturer", "house"},
    "name": {"name", "scent", "scent_name", "scent name", "perfume", "perfume_name", "product", "product_name", "title"},
    "concentration": {"concentration", "conc", "type", "strength", "edp edt", "edp_edt"},
    "size_ml": {"size", "size_ml", "ml", "volume", "volume_ml"},
    "audience": {"audience", "gender", "sex", "target"},
    "variant_type": {"variant_type", "variant type", "variant", "format"},
    "packaging": {"packaging", "package", "pack", "box", "presentation"},
    "is_tester": {"is_tester", "tester", "test"},
    "sku": {"sku", "our_sku", "our sku", "code", "item_code", "item code"},
    "ean": {"ean", "barcode", "bar_code", "upc"},
    "comments": {"comments", "comment", "notes", "note"},
    "collection_name": {"collection", "collection_name", "collection name", "subname", "sub_name", "sub name", "line", "perfume_line", "perfume line"},
    "release_year": {"release_year", "year", "launch_year"},
    "perfumer_name": {"perfumer", "perfumer_name", "nose"},
    "country_of_origin": {"brand_country", "country_of_origin", "origin"},
    "country_of_manufacture": {"made_in", "country_of_manufacture", "manufacture_country"},
}

CONCENTRATION_MAP = {
    "eau de parfum": "edp",
    "edp": "edp",
    "eau de toilette": "edt",
    "edt": "edt",
    "eau de cologne": "edc",
    "edc": "edc",
    "parfum": "parfum",
    "extrait": "extrait",
    "extrait de parfum": "extrait",
    "perfume oil": "perfume_oil",
}


@dataclass
class CatalogImportResult:
    rows_seen: int = 0
    rows_imported: int = 0
    brands_created: int = 0
    perfumes_created: int = 0
    perfumes_updated: int = 0
    variants_created: int = 0
    variants_updated: int = 0
    aliases_created: int = 0
    skipped_rows: list[str] = field(default_factory=list)


def _clean_header(value) -> str:
    return normalize_text(str(value or "")).replace(" ", "_")


def _column_key(header: str) -> str | None:
    header_text = normalize_text(header).replace("_", " ")
    for key, aliases in COLUMN_ALIASES.items():
        if header_text in {normalize_text(alias).replace("_", " ") for alias in aliases}:
            return key
    return None


def _text(row: dict, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value: str) -> Decimal | None:
    raw = (value or "").lower().replace("ml", "").replace(",", ".").strip()
    if not raw:
        return None
    try:
        return Decimal(raw).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _bool(value: str) -> bool:
    return normalize_text(value) in {"1", "yes", "true", "y", "tester", "test"}


def _concentration(value: str) -> str:
    return CONCENTRATION_MAP.get(normalize_text(value), normalize_text(value))


def _identity(value: str) -> str:
    return normalize_text(value)


def _variant_type_from_comments(value: str) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if "sample" in text or "semple" in text:
        return "sample_set" if "set" in text else "sample"
    if "travel" in text or "mini" in text:
        return "travel_set" if "set" in text else "travel"
    if "refill" in text:
        return "refill_set" if "set" in text else "refill"
    if "roll on" in text:
        return "roll_on"
    if "set" in text:
        return "set"
    return ""


def _read_csv(uploaded_file) -> list[dict]:
    content = uploaded_file.read().decode("utf-8-sig")
    reader = csv.DictReader(StringIO(content))
    return [{_column_key(header) or _clean_header(header): value for header, value in row.items()} for row in reader]


def _read_xlsx(uploaded_file) -> list[dict]:
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_column_key(str(header or "")) or _clean_header(header) for header in rows[0]]
    result = []
    for values in rows[1:]:
        result.append({header: value for header, value in zip(headers, values)})
    workbook.close()
    return result


def read_catalog_rows(uploaded_file) -> list[dict]:
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        return _read_csv(uploaded_file)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(uploaded_file)
    raise ValueError("Upload .xlsx or .csv catalogue files.")


@transaction.atomic
def import_catalog_file(uploaded_file, *, create_aliases: bool = True, update_existing: bool = True) -> CatalogImportResult:
    result = CatalogImportResult()
    rows = read_catalog_rows(uploaded_file)
    brand_cache = {_identity(brand.name): brand for brand in Brand.objects.all()}
    perfume_cache = {
        (perfume.brand_id, _identity(perfume.name), perfume.concentration or "", perfume.audience or ""): perfume
        for perfume in Perfume.objects.select_related("brand").all()
    }
    variant_cache = {
        (variant.perfume_id, variant.size_ml, variant.packaging or "", variant.variant_type or "", variant.is_tester): variant
        for variant in PerfumeVariant.objects.all()
    }
    brand_alias_cache = {
        (alias.brand_id, alias.supplier_id, _identity(alias.alias_text))
        for alias in BrandAlias.objects.all()
    }
    product_alias_cache = {
        (alias.brand_id, alias.perfume_id, alias.supplier_id, _identity(alias.alias_text))
        for alias in ProductAlias.objects.all()
    }

    for index, row in enumerate(rows, start=2):
        result.rows_seen += 1
        brand_name = _text(row, "brand")
        perfume_name = _text(row, "name")
        if not brand_name or not perfume_name:
            result.skipped_rows.append(f"Row {index}: missing brand or scent name")
            continue

        brand_key = _identity(brand_name)
        brand = brand_cache.get(brand_key)
        if not brand:
            brand = Brand.objects.create(name=brand_name, country_of_origin=_text(row, "country_of_origin"))
            brand_cache[brand_key] = brand
            result.brands_created += 1
        elif update_existing and _text(row, "country_of_origin") and not brand.country_of_origin:
            brand.country_of_origin = _text(row, "country_of_origin")
            brand.save(update_fields=["country_of_origin", "updated_at"])

        concentration = _concentration(_text(row, "concentration"))
        audience = normalize_text(_text(row, "audience"))
        comments = _text(row, "comments")
        collection_name = _text(row, "collection_name")
        perfume_key = (brand.id, _identity(perfume_name), concentration, audience)
        perfume = perfume_cache.get(perfume_key)
        if perfume:
            if update_existing:
                perfume.collection_name = collection_name or perfume.collection_name
                perfume.perfumer_name = _text(row, "perfumer_name") or perfume.perfumer_name
                perfume.country_of_manufacture = _text(row, "country_of_manufacture") or perfume.country_of_manufacture
                if _text(row, "release_year").isdigit():
                    perfume.release_year = int(_text(row, "release_year"))
                perfume.save()
                result.perfumes_updated += 1
        else:
            perfume = Perfume.objects.create(
                brand=brand,
                name=perfume_name,
                concentration=concentration,
                audience=audience,
                collection_name=collection_name,
                release_year=int(_text(row, "release_year")) if _text(row, "release_year").isdigit() else None,
                perfumer_name=_text(row, "perfumer_name"),
                country_of_manufacture=_text(row, "country_of_manufacture"),
            )
            perfume_cache[perfume_key] = perfume
            result.perfumes_created += 1

        size_ml = _decimal(_text(row, "size_ml"))
        raw_packaging = normalize_text(_text(row, "packaging"))
        variant_type = normalize_text(_text(row, "variant_type")) or _variant_type_from_comments(comments) or "standard"
        is_tester = (
            _bool(_text(row, "is_tester"))
            or raw_packaging == "tester"
            or "tester" in normalize_text(comments)
            or "tetser" in normalize_text(comments)
        )
        packaging = "" if raw_packaging == "tester" else raw_packaging
        if size_ml or packaging or variant_type or is_tester or _text(row, "sku") or _text(row, "ean"):
            variant_key = (perfume.id, size_ml, packaging, variant_type, is_tester)
            variant = variant_cache.get(variant_key)
            if not variant:
                variant = PerfumeVariant.objects.create(
                    perfume=perfume,
                    size_ml=size_ml,
                    packaging=packaging,
                    variant_type=variant_type,
                    is_tester=is_tester,
                    sku=_text(row, "sku"),
                    ean=_text(row, "ean"),
                )
                variant_cache[variant_key] = variant
                result.variants_created += 1
            elif update_existing:
                variant.sku = _text(row, "sku") or variant.sku
                variant.ean = _text(row, "ean") or variant.ean
                variant.save()
                result.variants_updated += 1

        if create_aliases:
            brand_alias_key = (brand.id, None, _identity(brand_name))
            if brand_alias_key not in brand_alias_cache:
                BrandAlias.objects.create(
                    brand=brand,
                    supplier=None,
                    alias_text=brand_name,
                    normalized_alias=normalize_text(brand_name),
                    priority=50,
                    active=True,
                )
                brand_alias_cache.add(brand_alias_key)
                result.aliases_created += 1
            product_alias_key = (brand.id, perfume.id, None, _identity(perfume_name))
            if product_alias_key not in product_alias_cache:
                ProductAlias.objects.create(
                    brand=brand,
                    perfume=perfume,
                    supplier=None,
                    alias_text=perfume_name,
                    canonical_text=perfume.name,
                    concentration=perfume.concentration,
                    audience=perfume.audience,
                    priority=50,
                    active=True,
                )
                product_alias_cache.add(product_alias_key)
                result.aliases_created += 1
        result.rows_imported += 1
    return result
